import os
import time
import base64
import logging
import requests
import pandas as pd
import gspread
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from dotenv import load_dotenv

# ==========================================
# 0. CARREGAMENTO BLINDADO DE CREDENCIAIS
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(BASE_DIR, '.env')
load_dotenv(ENV_PATH)

# ==========================================
# 0.1 CONFIGURAÇÃO DE LOGGING EM ARQUIVO
# ==========================================
LOG_PATH = os.path.join(BASE_DIR, 'logs', 'robo_sisab.log')
os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(LOG_PATH, encoding='utf-8'),
        logging.StreamHandler()  # Mantém saída no terminal também
    ]
)
log = logging.getLogger(__name__)

# ==========================================
# 1. CONFIGURAÇÕES FIXAS DO SISTEMA
# ==========================================
# ANO_FILTRO agora é dinâmico — muda automaticamente a cada ano
ANO_FILTRO = str(datetime.now().year)

URL_SITE = "https://relatorioaps.saude.gov.br/"
ID_PLANILHA_SHEETS   = os.getenv("ID_PLANILHA_SHEETS")
PASTA_EGESTOR        = os.getenv("PASTA_EGESTOR")         # Ex: \\SRV-FS\dics\GEPAP\{ANO}\IAF\...
NOME_ABA_SHEETS      = "e-Gestor"

# E-mail — destinatário dos avisos
EMAIL_DESTINATARIO = os.getenv("EMAIL_DESTINATARIO")

# Planilha local no servidor e eCIEGES
PLANILHA_LOCAL       = os.getenv("PLANILHA_LOCAL")        # Ex: \\SRV-FS\dics\GEPAP\arquivo.xlsx
ECIEGES_URL          = "http://ecieges.saude.df.gov.br/data-suite/formularios/responder/472"
ECIEGES_USUARIO      = os.getenv("ECIEGES_USUARIO")
ECIEGES_SENHA        = os.getenv("ECIEGES_SENHA")

# Scopes do Gmail API (adicionado ao token existente do Google)
GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/gmail.send",
]

# Resolve o caminho da pasta substituindo o ano dinamicamente
if PASTA_EGESTOR:
    PASTA_EGESTOR = PASTA_EGESTOR.replace("{ANO}", ANO_FILTRO)

# ==========================================
# 2. NOTIFICAÇÃO POR E-MAIL (Gmail API — HTTPS)
# ==========================================
def _get_gmail_service():
    """
    Reutiliza o token.json já existente no projeto.
    Se o token não tiver o scope do Gmail, apaga o token.json
    e na próxima execução o navegador pedirá nova autorização.
    """
    token_path = os.path.join(BASE_DIR, 'token.json')
    creds_path = os.path.join(BASE_DIR, 'credentials.json')

    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, GMAIL_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(GoogleRequest())
        else:
            from google_auth_oauthlib.flow import InstalledAppFlow
            flow = InstalledAppFlow.from_client_secrets_file(creds_path, GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'w') as token_file:
            token_file.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)


def enviar_email(assunto, corpo_html):
    """
    Envia e-mail via Gmail API (HTTPS porta 443).
    Funciona mesmo em redes que bloqueiam SMTP (465/587).
    Silencioso em caso de falha — não interrompe o fluxo principal.
    """
    if not EMAIL_DESTINATARIO:
        log.warning("Notificação por e-mail ignorada — EMAIL_DESTINATARIO ausente no .env.")
        return

    try:
        service = _get_gmail_service()

        msg = MIMEMultipart("alternative")
        msg["Subject"] = assunto
        msg["To"]      = EMAIL_DESTINATARIO
        msg.attach(MIMEText(corpo_html, "html", "utf-8"))

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(
            userId="me",
            body={"raw": raw}
        ).execute()

        log.info(f"E-mail enviado para {EMAIL_DESTINATARIO}: {assunto}")

    except Exception as e:
        log.warning(f"Falha ao enviar e-mail (não crítico): {e}")


def _corpo_base(titulo, cor_titulo, icone, linhas_corpo):
    """Monta o HTML padrão do e-mail com título e linhas de conteúdo."""
    itens_html = "".join(f"<li style='margin:6px 0'>{l}</li>" for l in linhas_corpo)
    agora = datetime.now().strftime("%d/%m/%Y às %H:%M")
    return f"""
    <div style="font-family:Calibri,Arial,sans-serif;max-width:600px;margin:0 auto;
                border:1px solid #ddd;border-radius:8px;overflow:hidden">
      <div style="background:{cor_titulo};padding:20px 24px">
        <h2 style="margin:0;color:#fff;font-size:20px">{icone} Robô SISAB — {titulo}</h2>
      </div>
      <div style="padding:24px;background:#fafafa">
        <ul style="padding-left:20px;color:#333;font-size:14px;line-height:1.6">
          {itens_html}
        </ul>
        <p style="margin-top:20px;font-size:12px;color:#999">
          Execução em {agora} &nbsp;|&nbsp; Ano de referência: {ANO_FILTRO}
        </p>
      </div>
    </div>"""


def notificar_sucesso(parcela):
    enviar_email(
        assunto=f"✅ [SISAB] Parcela {parcela} integrada com sucesso",
        corpo_html=_corpo_base(
            titulo="Execução concluída",
            cor_titulo="#1a7f4b",
            icone="✅",
            linhas_corpo=[
                f"<b>Parcela processada:</b> {parcela}",
                f"<b>Ano:</b> {ANO_FILTRO}",
                "Dados extraídos do portal e-Gestor APS.",
                "Planilha no Google Sheets atualizada.",
                "Planilha local no servidor sincronizada.",
                "Upload no eCIEGES realizado com sucesso.",
            ]
        )
    )


def notificar_parcela_indisponivel(parcela):
    enviar_email(
        assunto=f"⚠️ [SISAB] Parcela {parcela} ainda não disponível no site",
        corpo_html=_corpo_base(
            titulo="Parcela indisponível",
            cor_titulo="#e6a817",
            icone="⚠️",
            linhas_corpo=[
                f"<b>Parcela buscada:</b> {parcela}",
                f"<b>Ano:</b> {ANO_FILTRO}",
                "O portal e-Gestor APS ainda não publicou esta parcela.",
                "Nenhuma alteração foi feita na planilha consolidada nem no eCIEGES.",
                "O robô tentará novamente na próxima execução agendada.",
            ]
        )
    )


def notificar_erro(etapa, detalhe):
    enviar_email(
        assunto=f"❌ [SISAB] Erro na etapa: {etapa}",
        corpo_html=_corpo_base(
            titulo="Erro na execução",
            cor_titulo="#c0392b",
            icone="❌",
            linhas_corpo=[
                f"<b>Etapa com falha:</b> {etapa}",
                f"<b>Detalhe:</b> <code style='background:#f5f5f5;padding:2px 6px;"
                f"border-radius:4px;font-size:13px'>{detalhe}</code>",
                "Verifique o arquivo <b>logs/robo_sisab.log</b> para o rastreamento completo.",
            ]
        )
    )


# ==========================================
# 3. MOTOR DE INTELIGÊNCIA: DESCOBRIR PRÓXIMA PARCELA
# ==========================================
def descobrir_proxima_parcela():
    log.info("Ligando ao Google Sheets para descobrir a última parcela processada...")

    gc = gspread.oauth(
        credentials_filename=os.path.join(BASE_DIR, 'credentials.json'),
        authorized_user_filename=os.path.join(BASE_DIR, 'token.json'),
        scopes=GMAIL_SCOPES
    )

    planilha = gc.open_by_key(ID_PLANILHA_SHEETS)
    aba = planilha.worksheet(NOME_ABA_SHEETS)
    dados_no_sheets = aba.get_all_values()

    # Procura de baixo para cima a última parcela válida
    ultima_parcela = "0/12"
    for linha in reversed(dados_no_sheets):
        if len(linha) > 4 and "/" in str(linha[4]):
            ultima_parcela = str(linha[4]).strip()
            break

    log.info(f"Última parcela na planilha: {ultima_parcela}")

    # Extrai o número (ex: de "4/12" tira o 4) e soma 1
    numerador_atual = int(ultima_parcela.replace("'", "").split('/')[0])
    proximo_numerador = numerador_atual + 1

    if proximo_numerador > 12:
        log.info("Todas as 12 parcelas deste ano já foram integradas! Nada a fazer.")
        return None, None

    # Parcela para busca no site (sem zero à esquerda — ex: 6/12)
    parcela_site = f"{proximo_numerador}/12"
    # Parcela para gravar no Sheets (com zero à esquerda — ex: 06/12)
    parcela_sheets = f"{proximo_numerador:02d}/12"

    log.info(f"Alvo automático definido para: {parcela_sheets}")

    # Nome de arquivo dinâmico sem precisar saber o mês de cabeça
    nome_arquivo = f"Relatorio_Parcela_{parcela_sheets.replace('/', 'de')}_{ANO_FILTRO}.xlsx"
    caminho_completo = os.path.join(PASTA_EGESTOR, nome_arquivo)

    return parcela_site, parcela_sheets, caminho_completo

# ==========================================
# 4. EXTRAÇÃO DE DADOS (PLAYWRIGHT)
# ==========================================
def baixar_relatorio_egestor(parcela_site, parcela_sheets, caminho_completo):
    log.info(f"Iniciando o robô para baixar a parcela {parcela_sheets}...")

    with sync_playwright() as p:
        # headless=True para rodar em background (Agendador de Tarefas, servidor sem desktop)
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        log.info("Acessando a página inicial...")
        page.goto(URL_SITE)
        page.wait_for_load_state("networkidle")

        log.info("Navegando para Financiamento APS...")
        page.click("a[href='/gerenciaaps/pagamento']")
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(2000)

        log.info("Preenchendo filtros...")
        page.click("span#tipo-unidade")
        page.wait_for_timeout(1000)
        page.click("li:has-text('Município')")
        page.wait_for_timeout(2000)

        page.click("span#estado")
        page.wait_for_timeout(1000)
        page.click("li:has-text('DISTRITO FEDERAL')")
        page.wait_for_timeout(3000)

        page.click("span#municipio")
        page.wait_for_timeout(1000)
        page.locator("li:has-text('BRASÍLIA'):visible").first.click()

        page.click("span#ano")
        page.wait_for_timeout(1000)
        page.click(f"li:has-text('{ANO_FILTRO}')")

        # Parcela início e fim — busca no site sem zero à esquerda (ex: 6/12)
        try:
            page.click("span#parcela-inicio")
            page.wait_for_timeout(1000)
            page.locator(f"li:has-text('{parcela_site}'):visible").first.click(timeout=10000)
            page.wait_for_timeout(2000)

            page.click("span#parcela-fim")
            page.wait_for_timeout(1000)
            page.locator(f"li:has-text('{parcela_site}'):visible").first.click(timeout=10000)

        except Exception:
            log.warning(f"Parcela {parcela_sheets} ainda não está disponível no site. Encerrando.")
            browser.close()
            notificar_parcela_indisponivel(parcela_sheets)
            return None

        log.info("Executando consulta...")
        page.click("button[aria-label='Ver em tela']")
        page.wait_for_timeout(3000)

        log.info("Expandindo primeira camada da árvore...")
        page.locator("span.pi-chevron-right").first.click()
        page.wait_for_timeout(2500)

        log.info("Verificando estruturas históricas e atuais da página...")
        try:
            texto_antigo = page.get_by_text(
                "Demais programas, serviços e equipes da Atenção Primária à Saúde", exact=True
            )
            texto_antigo.wait_for(state="visible", timeout=5000)
            texto_antigo.locator("xpath=ancestor::tr[1]").locator("button").first.click()
            page.wait_for_timeout(1500)
        except Exception:
            pass

        try:
            texto_novo = page.get_by_text(
                "Incentivo financeiro da APS - Promoção à saúde", exact=True
            )
            texto_novo.wait_for(state="visible", timeout=5000)
            texto_novo.locator("xpath=ancestor::tr[1]").locator("button").first.click()
            page.wait_for_timeout(1500)
        except Exception:
            pass

        log.info("Abrindo detalhes em nova aba...")
        texto_atividade = page.get_by_text("Incentivo de Atividade Física", exact=True)

        with context.expect_page() as new_page_info:
            texto_atividade.locator(
                "xpath=ancestor::tr[1]"
            ).locator("button:has-text('Ver Detalhes')").click()

        aba_relatorio = new_page_info.value
        aba_relatorio.wait_for_load_state("networkidle")
        log.info("Foco alterado para a aba do relatório.")

        sucesso_download = False
        try:
            log.info("Iniciando download do Excel...")
            botao_dl = aba_relatorio.locator("button[aria-label='Download do excel']").first

            with aba_relatorio.expect_download(timeout=60000) as download_info:
                botao_dl.click(force=True)

            download = download_info.value
            download.save_as(caminho_completo)
            log.info(f"Download concluído: {caminho_completo}")
            sucesso_download = True

        except Exception as e:
            log.error(f"Erro ao tentar baixar o arquivo: {e}")
            notificar_erro("Download do relatório e-Gestor", str(e))

        browser.close()
        return caminho_completo if sucesso_download else None


# ==========================================
# 4.1 WRAPPER COM RETRY PARA TIMEOUTS DO SITE
# ==========================================
def baixar_com_retry(parcela_site, parcela_sheets, caminho_completo,
                     max_tentativas=3, espera_segundos=300):
    """
    Encapsula `baixar_relatorio_egestor` com tentativas automáticas
    em caso de timeout do site do governo (instabilidade temporária).

    - max_tentativas: 3 (configurável)
    - espera_segundos: 300s = 5 minutos entre tentativas
    - Só refaz em PlaywrightTimeoutError. Outros erros são propagados.
    - Parcela indisponível NÃO é timeout, então não dispara retry.
    """
    for tentativa in range(1, max_tentativas + 1):
        try:
            log.info(f"Tentativa {tentativa}/{max_tentativas} de baixar parcela {parcela_sheets}...")
            resultado = baixar_relatorio_egestor(parcela_site, parcela_sheets, caminho_completo)
            return resultado  # Sucesso ou parcela indisponível → encerra sem retry

        except PlaywrightTimeoutError as e:
            log.warning(f"Timeout na tentativa {tentativa}: {e}")

            if tentativa < max_tentativas:
                minutos = espera_segundos // 60
                log.info(f"Aguardando {minutos} minutos antes da próxima tentativa...")
                time.sleep(espera_segundos)
            else:
                log.error(f"Todas as {max_tentativas} tentativas falharam por timeout. Desistindo.")
                notificar_erro(
                    "Timeout no site do e-Gestor",
                    f"Tentei {max_tentativas} vezes baixar a parcela {parcela_sheets} "
                    f"com intervalo de {espera_segundos // 60} minutos entre tentativas, "
                    f"mas o site continua instável. Último erro: {e}"
                )
                return None

    return None

# ==========================================
# 5. INTEGRAÇÃO E FORMATAÇÃO (GOOGLE SHEETS)
# ==========================================
def enviar_para_sheets(caminho_arquivo, nome_aba, parcela_alvo):
    log.info(f"Iniciando integração para a aba: {nome_aba}...")

    df = pd.read_excel(caminho_arquivo)

    if nome_aba == "e-Gestor":
        try:
            df = df.drop(df.columns[[13, 14]], axis=1)
        except Exception:
            pass

    df = df.fillna("")

    # --------------------------------------------------
    # FILTRO DE HIGIENIZAÇÃO E BLOQUEIO DE DATA
    # --------------------------------------------------
    valores_para_subir = []
    for linha in df.values.tolist():
        linha_limpa = []
        for i, celula in enumerate(linha):
            valor = str(celula).strip()

            # Coluna E (índice 4): injeta aspa simples para bloquear conversão de data no Sheets
            if i == 4 and "/" in valor:
                try:
                    numerador, denominador = valor.split('/')
                    valor = f"'{int(numerador):02d}/{denominador}"
                except Exception:
                    pass

            linha_limpa.append(valor)
        valores_para_subir.append(linha_limpa)
    # --------------------------------------------------

    try:
        gc = gspread.oauth(
            credentials_filename=os.path.join(BASE_DIR, 'credentials.json'),
            authorized_user_filename=os.path.join(BASE_DIR, 'token.json'),
            scopes=GMAIL_SCOPES
        )

        planilha = gc.open_by_key(ID_PLANILHA_SHEETS)
        aba = planilha.worksheet(nome_aba)

        log.info("Verificando antiduplicidade...")
        dados_no_sheets = aba.get_all_values()

        # Normaliza o alvo para comparação justa (sem zeros, sem aspas, minúsculas)
        parcela_alvo_limpa = parcela_alvo.strip().lower().replace("'", "").lstrip("0")

        ja_existe = False
        for linha in dados_no_sheets:
            if len(linha) > 4:
                coluna_parcela = str(linha[4]).strip().lower().replace("'", "").lstrip("0")
                if coluna_parcela == parcela_alvo_limpa:
                    ja_existe = True
                    break

        if ja_existe:
            log.warning(f"Parcela {parcela_alvo} já existe na planilha. Envio abortado para evitar duplicidade.")
            return False
        else:
            log.info(f"Adicionando dados da parcela {parcela_alvo} ao final da tabela...")
            aba.append_rows(valores_para_subir, value_input_option="USER_ENTERED")
            log.info("Dados integrados com sucesso.")

            log.info("Aplicando formatação Calibri 10...")
            total_colunas = len(df.columns)
            letra_final = gspread.utils.rowcol_to_a1(1, total_colunas).replace("1", "")
            intervalo_formatacao = f"A:{letra_final}"

            aba.format(intervalo_formatacao, {
                "textFormat": {
                    "fontFamily": "Calibri",
                    "fontSize": 10
                }
            })
            log.info("Formatação visual aplicada com sucesso.")
            return True

    except Exception as e:
        log.error(f"Erro na autenticação ou envio para o Sheets: {e}")
        notificar_erro("Envio para o Google Sheets", str(e))
        return False

# ==========================================
# 6. SHEETS → PLANILHA LOCAL NO SERVIDOR
# ==========================================
def atualizar_planilha_local():
    """
    Baixa todos os dados da aba 'e-Gestor' do Google Sheets e substitui
    a aba 'egestor' na planilha local do servidor (truncate + insert).
    """
    log.info("Iniciando sincronização: Google Sheets → Planilha local...")

    if not PLANILHA_LOCAL:
        log.warning("PLANILHA_LOCAL não definida no .env — etapa ignorada.")
        return False

    if not os.path.exists(PLANILHA_LOCAL):
        log.error(f"Planilha local não encontrada: {PLANILHA_LOCAL}")
        notificar_erro("Planilha local", f"Arquivo não encontrado: {PLANILHA_LOCAL}")
        return False

    try:
        # Baixa dados completos do Sheets
        gc = gspread.oauth(
            credentials_filename=os.path.join(BASE_DIR, 'credentials.json'),
            authorized_user_filename=os.path.join(BASE_DIR, 'token.json'),
            scopes=GMAIL_SCOPES
        )
        planilha = gc.open_by_key(ID_PLANILHA_SHEETS)
        aba_sheets = planilha.worksheet(NOME_ABA_SHEETS)
        dados = aba_sheets.get_all_values()

        if not dados or len(dados) < 2:
            log.warning("Sheets sem dados suficientes para sincronizar.")
            return False

        # Monta DataFrame a partir do Sheets (1ª linha = cabeçalho)
        df_sheets = pd.DataFrame(dados[1:], columns=dados[0])

        # Remove aspas simples de proteção de data inseridas anteriormente
        df_sheets = df_sheets.map(
            lambda x: x.lstrip("'") if isinstance(x, str) else x
        )

        log.info(f"Dados baixados do Sheets: {len(df_sheets)} linhas.")

        # Abre a planilha local preservando as outras abas
        from openpyxl import load_workbook
        wb = load_workbook(PLANILHA_LOCAL)

        # Remove a aba e-Gestor antiga e cria uma nova zerada (truncate)
        if "e-Gestor" in wb.sheetnames:
            del wb["e-Gestor"]

        ws = wb.create_sheet("e-Gestor")

        # Escreve cabeçalho
        for col_idx, col_name in enumerate(df_sheets.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Escreve dados linha a linha (insert)
        for row_idx, row in enumerate(df_sheets.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(PLANILHA_LOCAL)
        log.info(f"Planilha local atualizada com sucesso: {PLANILHA_LOCAL}")
        return True

    except Exception as e:
        log.error(f"Erro ao atualizar planilha local: {e}")
        notificar_erro("Sincronização planilha local", str(e))
        return False


# ==========================================
# 7. UPLOAD NO eCIEGES
# ==========================================
def upload_ecieges():
    """
    Faz login no eCIEGES via Playwright e realiza o upload
    da planilha local atualizada na tabela egestor (form 472).
    """
    log.info("Iniciando upload no eCIEGES...")

    if not all([ECIEGES_USUARIO, ECIEGES_SENHA]):
        log.warning("Credenciais do eCIEGES ausentes no .env — etapa ignorada.")
        return False

    if not PLANILHA_LOCAL or not os.path.exists(PLANILHA_LOCAL):
        log.error("Planilha local não encontrada para upload no eCIEGES.")
        return False

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page    = context.new_page()

        try:
            # Passo 1: Login
            log.info("Acessando página de login do eCIEGES...")
            page.goto("http://ecieges.saude.df.gov.br/login", wait_until="networkidle")

            # Seletores específicos do eCIEGES (Material UI / React)
            page.fill("input[name='login']", ECIEGES_USUARIO)
            page.fill("input[name='password']", ECIEGES_SENHA)
            page.click("button[type='submit']:has-text('Login')")
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(3000)

            # Passo 2: Navega direto para o formulário 472 (e-Gestor)
            log.info("Navegando para o formulário e-Gestor (472)...")
            page.goto(ECIEGES_URL, wait_until="networkidle")
            page.wait_for_timeout(3000)

            # Passo 3: Clica no botão "Importar arquivo"
            # (página carrega via React e o botão fica fora da viewport — precisa scroll)
            log.info("Aguardando o botão 'Importar arquivo' aparecer no DOM...")
            botao_importar = page.locator("button:has-text('Importar arquivo')").first

            try:
                botao_importar.wait_for(state="attached", timeout=60000)
            except Exception:
                # Captura screenshot e HTML para diagnóstico
                screenshot_path = os.path.join(BASE_DIR, 'logs', 'erro_ecieges.png')
                html_path       = os.path.join(BASE_DIR, 'logs', 'erro_ecieges.html')
                page.screenshot(path=screenshot_path, full_page=True)
                with open(html_path, 'w', encoding='utf-8') as f:
                    f.write(page.content())
                log.error(f"Botão não encontrado. Screenshot salvo em: {screenshot_path}")
                log.error(f"HTML da página salvo em: {html_path}")
                raise

            log.info("Rolando até o botão...")
            botao_importar.scroll_into_view_if_needed()
            page.wait_for_timeout(1500)

            log.info("Clicando em 'Importar arquivo'...")
            botao_importar.click()
            page.wait_for_timeout(2000)

            # Passo 3.1: Modal de confirmação "Importação de Planilha" (truncate)
            log.info("Aguardando modal de confirmação 'Importação de Planilha'...")
            page.locator("text=Importação de Planilha").wait_for(state="visible", timeout=15000)
            log.info("Confirmando importação (clicando em 'Sim')...")
            page.locator("button:has-text('Sim')").first.click()
            page.wait_for_timeout(2000)

            # Passo 4: Seleciona o arquivo XLSX
            log.info(f"Selecionando arquivo: {PLANILHA_LOCAL}")
            input_file = page.locator("input[type='file']").first
            input_file.wait_for(state="attached", timeout=10000)
            input_file.set_input_files(PLANILHA_LOCAL)
            page.wait_for_timeout(3000)

            # Passo 5: Modal "Selecione a aba da planilha" → escolher e-Gestor
            log.info("Aguardando modal de seleção de aba...")
            page.locator("text=Selecione a aba da planilha").wait_for(state="visible", timeout=15000)

            log.info("Abrindo dropdown de abas...")
            # MUI Select usa div[role='button'] com aria-haspopup='listbox'
            page.locator("[aria-haspopup='listbox']").last.click()
            page.wait_for_timeout(1500)

            log.info("Selecionando aba 'e-Gestor'...")
            page.locator("li:has-text('e-Gestor')").first.click()
            page.wait_for_timeout(1500)

            # Passo 6: Confirma com o botão "Importar essa aba"
            log.info("Clicando em 'Importar essa aba'...")
            page.locator("button:has-text('Importar essa aba')").first.click()

            # Aguarda processamento (upload de arquivo grande pode demorar)
            log.info("Aguardando processamento do upload...")
            page.wait_for_load_state("networkidle", timeout=120000)
            page.wait_for_timeout(5000)

            log.info("Upload no eCIEGES concluído com sucesso.")
            browser.close()
            return True

        except Exception as e:
            log.error(f"Erro durante upload no eCIEGES: {e}")
            notificar_erro("Upload eCIEGES", str(e))
            browser.close()
            return False


# ==========================================
# 8. ORQUESTRAÇÃO FINAL
# ==========================================
if __name__ == "__main__":
    log.info("=" * 60)
    log.info(f"ROBÔ SISAB INICIADO — Ano de referência: {ANO_FILTRO}")
    log.info("=" * 60)

    try:
        # Passo 1: Descobre qual parcela processar
        parcela_site, parcela_sheets, caminho_completo = descobrir_proxima_parcela()

        if not parcela_site:
            log.info("Nenhuma ação necessária. Encerrando.")
        else:
            # Passo 2: Baixa o relatório do portal e-Gestor com retry automático
            # (até 3 tentativas com 5 minutos de intervalo em caso de timeout do site)
            arquivo_egestor = baixar_com_retry(parcela_site, parcela_sheets, caminho_completo)

            if not arquivo_egestor:
                log.error("Envio para o Google Sheets cancelado — download falhou.")
            else:
                # Passo 3: Envia e consolida no Google Sheets (grava com zero: 06/12)
                sheets_ok = enviar_para_sheets(arquivo_egestor, NOME_ABA_SHEETS, parcela_sheets)

                if not sheets_ok:
                    log.warning("Etapas seguintes ignoradas pois o Sheets não foi atualizado.")
                else:
                    # Passo 4: Sheets → planilha local no servidor
                    local_ok = atualizar_planilha_local()

                    # Passo 5: Upload da planilha local no eCIEGES
                    ecieges_ok = upload_ecieges() if local_ok else False
                    if not local_ok:
                        log.warning("Upload no eCIEGES ignorado pois a planilha local não foi atualizada.")

                    # Passo 6: Notifica resultado completo
                    if local_ok and ecieges_ok:
                        notificar_sucesso(parcela_sheets)
                    else:
                        falhas = []
                        if not local_ok:
                            falhas.append("Sincronização da planilha local")
                        if not ecieges_ok:
                            falhas.append("Upload no eCIEGES")
                        notificar_erro(
                            "Etapas com falha após atualização do Sheets",
                            f"Parcela {parcela_sheets} gravada no Sheets, mas falhou em: "
                            + ", ".join(falhas) + ". Verifique o log."
                        )

    except Exception as e:
        log.exception(f"Erro crítico na execução geral: {e}")
        notificar_erro("Execução geral (erro crítico)", str(e))

    log.info("=" * 60)
    log.info("ROBÔ SISAB FINALIZADO")
    log.info("=" * 60)